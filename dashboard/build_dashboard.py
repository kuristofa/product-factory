"""Builder for the Phase 5 draft dashboard template (regenerable source of truth).

Produces a 4-tab .xlsx: Overview (product config + KPI scorecard), Content Log (the
data source), Funnel (conversion rates), Feedback (support + ratings). Seeded with
example data for the spec's example product. All KPIs/rates are Excel formulas, so the
sheet stays live: replace the example rows with real data and everything recalculates.

Regenerate:  python dashboard/build_dashboard.py
Then recalc:  python /mnt/skills/public/xlsx/scripts/recalc.py <file>
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent / "FascinateCopy_Product_Dashboard_TEMPLATE.xlsx"

# palette
ACCENT = "2F6F6B"
ACCENT_DK = "234F4C"
BAND = "EAF0F0"
INPUT_BLUE = "0000FF"
GREY = "5B6B7B"

FONT = "Arial"
thin = Side(style="thin", color="D6DDE3")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def hdr(cell):
    cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=ACCENT)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = BORDER


def label(cell, bold=True):
    cell.font = Font(name=FONT, bold=bold, size=10, color="1B2733")
    cell.fill = PatternFill("solid", fgColor=BAND)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER


def val(cell, num=None, blue=False, align="left"):
    cell.font = Font(name=FONT, size=10, color=INPUT_BLUE if blue else "1B2733")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = BORDER
    if num:
        cell.number_format = num


def section(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=ACCENT_DK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=ACCENT_DK)


wb = Workbook()

# ── CONTENT LOG (data source) ────────────────────────────────────────────────
cl = wb.active
cl.title = "Content Log"
cl.sheet_view.showGridLines = False
cl_headers = ["Date", "Content Type", "Title / Hook", "Keyword", "Reel Views",
              "Carousel Engagement", "DMs Generated", "Keyword Triggers",
              "Checkout Clicks", "Purchases", "Order-Bump Purchases", "Revenue"]
title = cl.cell(row=1, column=1, value="Content Log — example data (replace with real published content)")
title.font = Font(name=FONT, bold=True, size=12, color=ACCENT_DK)
cl.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cl_headers))
for j, h in enumerate(cl_headers, 1):
    hdr(cl.cell(row=3, column=j, value=h))

rows = [
    ["2025-06-02", "Reel", "The 20-min posting mistake", "REELS", 8200, 0, 210, 220, 74, 15, 5],
    ["2025-06-04", "Carousel", "5 hooks that book calls", "HOOKS", 0, 640, 95, 100, 40, 9, 3],
    ["2025-06-06", "Reel", "Why your Reels don't convert", "CONVERT", 6100, 0, 150, 160, 52, 11, 4],
    ["2025-06-09", "Post", "Client booked 5 calls in a week", "PROOF", 0, 210, 60, 62, 22, 6, 2],
    ["2025-06-11", "Reel", "Steal my keyword-to-DM setup", "SETUP", 9400, 0, 240, 250, 61, 12, 3],
    ["2025-06-13", "Carousel", "The DM script I use", "SCRIPT", 0, 520, 80, 85, 28, 5, 1],
    ["2025-06-16", "Reel", "Post this when you're busy", "BUSY", 5200, 0, 120, 130, 33, 2, 1],
    ["2025-06-18", "Post", "No time? 20 minutes is enough", "TIME", 0, 180, 45, 48, 15, 1, 1],
    ["2025-06-20", "Reel", "3 hooks for coaches", "COACH", 7300, 0, 190, 200, 40, 1, 0],
    ["2025-06-23", "Story", "Poll: what stops you posting?", "POLL", 0, 90, 30, 32, 8, 0, 0],
]
first_data = 4
for i, r in enumerate(rows):
    rr = first_data + i
    for j, v in enumerate(r, 1):
        val(cl.cell(row=rr, column=j, value=v),
            num="#,##0" if j >= 5 else None, align="right" if j >= 5 else "left")
    # Revenue formula: purchases*price + bump*bump_price (prices live on Overview)
    rev = cl.cell(row=rr, column=12,
                  value=f"=J{rr}*Overview!$B$7+K{rr}*Overview!$B$8")
    val(rev, num='$#,##0', align="right")
last_data = first_data + len(rows) - 1
total_row = last_data + 1
label(cl.cell(row=total_row, column=1, value="TOTALS"))
for j in range(2, 5):
    label(cl.cell(row=total_row, column=j, value=""), bold=False)
for j in range(5, 13):
    col = get_column_letter(j)
    c = cl.cell(row=total_row, column=j, value=f"=SUM({col}{first_data}:{col}{last_data})")
    c.font = Font(name=FONT, bold=True, size=10, color="1B2733")
    c.fill = PatternFill("solid", fgColor=BAND)
    c.border = BORDER
    c.number_format = '$#,##0' if j == 12 else "#,##0"
    c.alignment = Alignment(horizontal="right")
widths = [11, 13, 30, 11, 11, 13, 12, 13, 12, 11, 14, 12]
for j, w in enumerate(widths, 1):
    cl.column_dimensions[get_column_letter(j)].width = w
cl.freeze_panes = "A4"

# ── OVERVIEW (config + scorecard) ────────────────────────────────────────────
ov = wb.create_sheet("Overview", 0)
ov.sheet_view.showGridLines = False
t = ov.cell(row=1, column=1, value="Product Performance Dashboard")
t.font = Font(name=FONT, bold=True, size=16, color=ACCENT_DK)
st = ov.cell(row=2, column=1, value="Draft template · example data · replace inputs (blue) and the Content Log with real numbers")
st.font = Font(name=FONT, italic=True, size=9.5, color=GREY)

section(ov, 4, "PRODUCT CONFIGURATION", 2)
config = [
    ("Product", "The 20-Minute Reels System", False, None),
    ("Avatar", "Time-poor coaches with <5k followers", False, None),
    ("Core product price ($)", 27, True, '$#,##0'),
    ("Order bump price ($)", 9, True, '$#,##0'),
    ("Core deliverables", "Reel hook framework; Keyword-to-DM setup", False, None),
    ("Bonuses", "30 done-for-you hooks", False, None),
    ("Order bump", "Caption pack", False, None),
]
r = 5
for lab, v, blue, num in config:
    label(ov.cell(row=r, column=1, value=lab))
    val(ov.cell(row=r, column=2, value=v), num=num, blue=blue,
        align="right" if num else "left")
    r += 1

score_start = r + 1
section(ov, score_start, "KPI SCORECARD  (live formulas — pull from Content Log & Feedback)", 2)
kpis = [
    ("Content published", f"=COUNTA('Content Log'!C{first_data}:C{last_data})", "#,##0"),
    ("Total reel views", f"='Content Log'!E{total_row}", "#,##0"),
    ("Total carousel engagement", f"='Content Log'!F{total_row}", "#,##0"),
    ("Total DMs", f"='Content Log'!G{total_row}", "#,##0"),
    ("Keyword triggers", f"='Content Log'!H{total_row}", "#,##0"),
    ("Checkout clicks", f"='Content Log'!I{total_row}", "#,##0"),
    ("Purchases", f"='Content Log'!J{total_row}", "#,##0"),
    ("Order-bump purchases", f"='Content Log'!K{total_row}", "#,##0"),
    ("Revenue", f"='Content Log'!L{total_row}", '$#,##0'),
    ("Refunds (input)", 2, "#,##0"),                       # blue input
    ("Refund rate", None, "0.0%"),                          # formula set below
    ("Support issues", "=COUNTA(Feedback!C5:C8)", "#,##0"),
    ("Avg customer rating", "=IFERROR(AVERAGE(Feedback!C12:C17),0)", "0.0"),
]
kr = score_start + 1
purchases_row = None
refunds_row = None
for lab, formula, num in kpis:
    label(ov.cell(row=kr, column=1, value=lab))
    if lab == "Purchases":
        purchases_row = kr
    if lab == "Refunds (input)":
        refunds_row = kr
        val(ov.cell(row=kr, column=2, value=formula), num=num, blue=True, align="right")
    elif lab == "Refund rate":
        f = f"=IF(B{purchases_row}=0,0,B{refunds_row}/B{purchases_row})"
        val(ov.cell(row=kr, column=2, value=f), num=num, align="right")
    else:
        val(ov.cell(row=kr, column=2, value=formula), num=num, align="right")
    kr += 1
ov.column_dimensions["A"].width = 30
ov.column_dimensions["B"].width = 40

# ── FUNNEL ───────────────────────────────────────────────────────────────────
fn = wb.create_sheet("Funnel")
fn.sheet_view.showGridLines = False
t = fn.cell(row=1, column=1, value="Funnel — example data")
t.font = Font(name=FONT, bold=True, size=12, color=ACCENT_DK)
for j, h in enumerate(["Stage", "Count", "Conversion"], 1):
    hdr(fn.cell(row=3, column=j, value=h))
# reference Overview scorecard rows
ov_rv, ov_kt, ov_dm = f"Overview!B{score_start+2}", f"Overview!B{score_start+5}", f"Overview!B{score_start+4}"
ov_cc, ov_pu, ov_ob = f"Overview!B{score_start+6}", f"Overview!B{score_start+7}", f"Overview!B{score_start+8}"
ov_rr = f"Overview!B{score_start+11}"
funnel = [
    ("Reel views (reach)", f"={ov_rv}", None),
    ("Keyword triggers", f"={ov_kt}", "=IF(B4=0,0,B5/B4)"),
    ("DMs entered", f"={ov_dm}", "=IF(B5=0,0,B6/B5)"),
    ("Checkout clicks", f"={ov_cc}", "=IF(B6=0,0,B7/B6)"),
    ("Purchases", f"={ov_pu}", "=IF(B7=0,0,B8/B7)"),
    ("Order-bump attach", f"={ov_ob}", "=IF(B8=0,0,B9/B8)"),
    ("Refund rate", f"={ov_rr}", None),
]
fr = 4
for lab, count_f, conv_f in funnel:
    label(fn.cell(row=fr, column=1, value=lab))
    val(fn.cell(row=fr, column=2, value=count_f),
        num='0.0%' if lab == "Refund rate" else "#,##0", align="right")
    if conv_f:
        val(fn.cell(row=fr, column=3, value=conv_f), num="0.0%", align="right")
    else:
        val(fn.cell(row=fr, column=3, value="—"), align="center")
    fr += 1
fn.column_dimensions["A"].width = 24
fn.column_dimensions["B"].width = 14
fn.column_dimensions["C"].width = 14

# ── FEEDBACK ─────────────────────────────────────────────────────────────────
fb = wb.create_sheet("Feedback")
fb.sheet_view.showGridLines = False
t = fb.cell(row=1, column=1, value="Feedback & Support — example data")
t.font = Font(name=FONT, bold=True, size=12, color=ACCENT_DK)
section(fb, 3, "SUPPORT ISSUES", 4)
for j, h in enumerate(["Date", "Buyer", "Issue", "Status"], 1):
    hdr(fb.cell(row=4, column=j, value=h))
support = [
    ["2025-06-12", "buyer #14", "Couldn't access the Drive folder", "Resolved"],
    ["2025-06-19", "buyer #33", "Wanted a refund — changed mind, kept it", "Resolved"],
    ["2025-06-24", "buyer #51", "Asked how to set up the DM keyword", "Resolved"],
]
for i, r_ in enumerate(support):
    for j, v in enumerate(r_, 1):
        val(fb.cell(row=5 + i, column=j, value=v))
section(fb, 10, "CUSTOMER FEEDBACK", 4)
for j, h in enumerate(["Date", "Buyer", "Rating (1-5)", "Comment"], 1):
    hdr(fb.cell(row=11, column=j, value=h))
feedback = [
    ["2025-06-08", "buyer #6", 5, "Booked 3 calls in the first week."],
    ["2025-06-10", "buyer #9", 5, "The keyword setup alone was worth it."],
    ["2025-06-15", "buyer #22", 4, "Great, wanted more hook examples."],
    ["2025-06-17", "buyer #27", 5, "Finally something I can do in 20 min."],
    ["2025-06-21", "buyer #40", 3, "Good but setup took me a while."],
    ["2025-06-22", "buyer #44", 4, "Solid framework, clear delivery."],
]
for i, r_ in enumerate(feedback):
    for j, v in enumerate(r_, 1):
        val(fb.cell(row=12 + i, column=j, value=v),
            num="0" if j == 3 else None, align="center" if j == 3 else "left")
for col, w in zip("ABCD", [11, 12, 14, 44]):
    fb.column_dimensions[col].width = w

wb.save(OUT)
print(f"Wrote {OUT}")
